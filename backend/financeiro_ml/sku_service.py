"""CRUD da tabela sku_financeiro + import Excel."""
from decimal import Decimal
from datetime import datetime
from typing import BinaryIO

from openpyxl import load_workbook

from database import SessionLocal
from financeiro_ml.models import SkuFinanceiro


def upsert_sku(sku: str, *, custo_unit: Decimal, imposto_pct: Decimal, updated_by_id: int) -> None:
    with SessionLocal() as session:
        row = session.query(SkuFinanceiro).filter_by(sku=sku).first()
        if row is None:
            row = SkuFinanceiro(sku=sku, custo_unit=custo_unit, imposto_pct=imposto_pct,
                                  updated_by=updated_by_id, updated_at=datetime.utcnow())
            session.add(row)
        else:
            row.custo_unit = custo_unit
            row.imposto_pct = imposto_pct
            row.updated_by = updated_by_id
            row.updated_at = datetime.utcnow()
        session.commit()


def list_skus(query: str | None = None) -> dict[str, dict]:
    with SessionLocal() as session:
        q = session.query(SkuFinanceiro)
        if query:
            q = q.filter(SkuFinanceiro.sku.like(f"%{query}%"))
        return {
            row.sku: {
                "custo_unit": row.custo_unit,
                "imposto_pct": row.imposto_pct,
                "updated_at": row.updated_at,
            }
            for row in q.all()
        }


def delete_sku(sku: str) -> bool:
    with SessionLocal() as session:
        row = session.query(SkuFinanceiro).filter_by(sku=sku).first()
        if row is None:
            return False
        session.delete(row)
        session.commit()
        return True


def import_excel(file: BinaryIO, *, updated_by_id: int) -> dict:
    """Importa planilha com colunas: sku, custo_unit, imposto_pct.
    Aceita ordem qualquer das colunas (lê pelo header).
    Retorna {created, updated, errors: list[{linha, motivo}]}.
    """
    wb = load_workbook(file, data_only=True, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {"created": 0, "updated": 0, "errors": [{"linha": 0, "motivo": "planilha vazia"}]}

    header = {str(c).strip().lower(): i for i, c in enumerate(rows[0]) if c is not None}
    required = ("sku", "custo_unit", "imposto_pct")
    missing = [h for h in required if h not in header]
    if missing:
        return {"created": 0, "updated": 0, "errors": [{"linha": 1, "motivo": f"colunas faltando: {missing}"}]}

    existing = set(list_skus().keys())
    created = 0
    updated = 0
    errors = []

    for line_no, row in enumerate(rows[1:], start=2):
        try:
            sku = str(row[header["sku"]]).strip() if row[header["sku"]] is not None else ""
            if not sku:
                errors.append({"linha": line_no, "motivo": "sku vazio"})
                continue
            custo = Decimal(str(row[header["custo_unit"]] or "0"))
            imposto = Decimal(str(row[header["imposto_pct"]] or "0"))
            upsert_sku(sku, custo_unit=custo, imposto_pct=imposto, updated_by_id=updated_by_id)
            if sku in existing:
                updated += 1
            else:
                created += 1
                existing.add(sku)
        except Exception as e:
            errors.append({"linha": line_no, "motivo": str(e)})

    return {"created": created, "updated": updated, "errors": errors}
